# -*- coding: utf-8 -*-
"""data/*.csv から整形済みの .xlsx を生成する。

CSV を編集したら本スクリプトを再実行するだけで .xlsx が更新される。
配色は ../DESIGN.md のデザイントークンに準拠。

実行:  python3 build/build_xlsx.py
"""
import csv
import os

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.normpath(os.path.join(HERE, ".."))
DATA_DIR = os.path.join(ROOT, "data")
OUT_PATH = os.path.join(ROOT, "英単語_品詞別_意味グループ.xlsx")

# ── DESIGN.md デザイントークン ──────────────────────────────
INK = "1F2421"
PAPER = "F6F3EC"
SURFACE = "FFFDF8"
ACCENT = "2F7D6B"
ACCENT_SOFT = "E3EFE9"
GOLD = "B8893B"
LINE = "DDD6C8"

# CSV(表示名) → シート名／グループ列インデックス
SHEETS = [
    ("動詞.csv", "動詞", 1),
    ("名詞.csv", "名詞", 1),
    ("形容詞.csv", "形容詞", 1),
    ("副詞.csv", "副詞", 1),
    ("機能語.csv", "機能語", 1),
    ("機能別フレーズ.csv", "機能別フレーズ", 0),
]

COL_WIDTHS = {
    "言いたいこと": 24, "意味グループ": 15, "サブグループ": 14, "見出し語": 16,
    "コア語義": 18, "活用・派生": 20, "コロケーション・句動詞": 28,
    "例文(EN)": 32, "例文(JP)": 26, "頻度ランク": 8, "レジスター": 10, "メモ": 30,
    "機能": 15, "丁寧度・場面": 12, "トリガー(日)": 18, "表現(EN)": 28,
    "意味(JP)": 18, "例文": 30,
}
WRAP_COLS = {"コロケーション・句動詞", "例文(EN)", "例文(JP)", "メモ", "表現(EN)", "意味(JP)"}

thin = Side(style="thin", color=LINE)
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
HEADER_FILL = PatternFill("solid", fgColor=ACCENT)
GROUP_FILLS = [PatternFill("solid", fgColor=SURFACE), PatternFill("solid", fgColor=ACCENT_SOFT)]
HEADER_FONT = Font(name="Hiragino Sans", color="FFFFFF", bold=True, size=11)
BODY_FONT = Font(name="Hiragino Sans", color=INK, size=11)
GROUP_FONT = Font(name="Hiragino Sans", color=INK, bold=True, size=11)


def read_csv(path):
    with open(path, encoding="utf-8-sig", newline="") as f:
        rows = list(csv.reader(f))
    return rows[0], rows[1:]


def style_sheet(ws, headers, body, group_idx):
    # ヘッダー行
    ws.append(headers)
    for col, name in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col)
        c.fill, c.font, c.border = HEADER_FILL, HEADER_FONT, BORDER
        c.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col)].width = COL_WIDTHS.get(name, 16)
    ws.row_dimensions[1].height = 26

    # 本体：意味グループが変わるたびに背景色を切り替えてブロックを可視化
    prev_group, band = None, 0
    for r, row in enumerate(body, start=2):
        group = row[group_idx] if group_idx < len(row) else ""
        if group != prev_group:
            band ^= 1
            prev_group = group
        fill = GROUP_FILLS[band]
        for col, value in enumerate(row, start=1):
            c = ws.cell(row=r, column=col, value=value)
            c.fill, c.border = fill, BORDER
            wrap = headers[col - 1] in WRAP_COLS
            c.alignment = Alignment(vertical="top", wrap_text=wrap)
            c.font = GROUP_FONT if (col - 1) == group_idx else BODY_FONT

    ws.freeze_panes = "B2"  # 1行目＋「言いたいこと/機能」列を固定（逆引きキーを常時表示）
    last_col = get_column_letter(len(headers))
    ws.auto_filter.ref = f"A1:{last_col}{len(body) + 1}"
    return len(body)


def build_index_sheet(ws, counts):
    ws.sheet_properties.tabColor = GOLD
    ws["A1"] = "英単語 品詞別 × 意味グループ データベース"
    ws["A1"].font = Font(name="Hiragino Mincho ProN", bold=True, size=16, color=INK)
    ws["A2"] = "会話で「言いたいこと」を1手で英語に変換するための逆引き語彙集"
    ws["A2"].font = Font(name="Hiragino Sans", size=11, color="55605A")

    guide = [
        "",
        "■ 使い方（逆引き）",
        "  1. 品詞のタブを開く（動詞／名詞／形容詞／副詞／機能語／機能別フレーズ）",
        "  2. ヘッダーの▼フィルタで「意味グループ」を絞る（例: 動詞→「思考・意見」）",
        "  3. または『言いたいこと』列を Ctrl+F で日本語検索（例:「依頼」「〜になる」）",
        "  4. 『見出し語』『コロケーション・句動詞』『例文』を見て、そのまま英作文に使う",
        "",
        "■ 列の意味",
        "  言いたいこと … 日本語の発話意図（逆引きキー）",
        "  意味グループ／サブグループ … 何を伝えたいかによる分類（できるだけ細分化）",
        "  活用・派生 … 動詞は 過去/過去分詞/-ing、形容詞は 比較級/最上級 など",
        "  頻度ランク … COCA/NGSL 由来の概算順位（100+ は概略、機能語は英語全体順位）",
        "  レジスター … 日常／口語／ビジネス／書き言葉 の使用域",
        "",
        "■ シート一覧",
    ]
    for i, kv in enumerate(counts):
        guide.append(f"  ・{kv[0]}：{kv[1]} 語（行）")
    guide += [
        "",
        "■ 拡張のしかた（Phase 2: NGSLコア約800〜1000語へ）",
        "  data/ の各CSVに同じ列で行を追記 → `python3 build/build_xlsx.py` を再実行",
        "",
        "■ 分類の根拠 … 動詞=CEFR機能+Levin動詞クラス／形容詞=Dixon意味型／副詞・名詞=COCA意味フィールド",
        "  詳細は同フォルダの 分類表.md を参照。関連は 04_ESP教材 / 03_学習ルーティン とリンク。",
    ]
    for i, line in enumerate(guide, start=4):
        cell = ws.cell(row=i, column=1, value=line)
        if line.startswith("■"):
            cell.font = Font(name="Hiragino Sans", bold=True, size=12, color=ACCENT)
        else:
            cell.font = Font(name="Hiragino Sans", size=11, color=INK)
    ws.column_dimensions["A"].width = 90
    ws.sheet_view.showGridLines = False


def main():
    wb = Workbook()
    index_ws = wb.active
    index_ws.title = "目次・使い方"

    counts = []
    for filename, sheet_name, group_idx in SHEETS:
        path = os.path.join(DATA_DIR, filename)
        headers, body = read_csv(path)
        ws = wb.create_sheet(title=sheet_name)
        n = style_sheet(ws, headers, body, group_idx)
        counts.append((sheet_name, n))

    build_index_sheet(index_ws, counts)
    wb.save(OUT_PATH)

    total = sum(n for _, n in counts)
    print(f"✓ 生成: {OUT_PATH}")
    for name, n in counts:
        print(f"    {name:12s} {n:3d} 行")
    print(f"  合計 {total} 語（行）／ シート数 {len(wb.sheetnames)}")
    return OUT_PATH


if __name__ == "__main__":
    main()
